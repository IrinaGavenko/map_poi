#!/usr/bin/env python3
"""Convert FWC26 POI spreadsheet sheet `20260624` to app Point[] JSON."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print('Install openpyxl: pip install openpyxl', file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / 'src/data/260624-fwc26-fifa-interactive-maps-pois.xlsx'
DEFAULT_OUT = ROOT / 'src/data/points.json'

# Excel `icon` column -> app point category ids (see src/types/categories.ts)
ICON_TO_CATEGORY = {
    'airport': 'airport',
    'car': 'car_rental',
    'fanfestival': 'fifa_fan_festival',
    'fifastore': 'fifa_store',
    'gate': 'gate',
    'hotel': 'hotel',
    'rewardslandmark': 'fifa_reward',
    'fifamuseum': 'fifa_museum',
    'rewardsmuseum': 'fifa_museum',
    'parking': 'parking_lot',
    'paidparking': 'parking_lot',
    'rideshare': 'shuttle_stop',
    'shuttle': 'shuttle_stop',
    'transport': 'transit_stop',
    'ferry': 'transit_station',
    'taxi': 'transit_stop',
    'bikeshare': 'transit_stop',
    'bikeparking': 'parking_lot',
}


def slugify(value: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', value.lower()).strip('-')
    return slug or 'poi'


def as_bool(value) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    return str(value).strip().lower() in {'true', 'yes', '1'}


def clean_text(value) -> str:
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def to_picture(img) -> list[str]:
    path = clean_text(img)
    return [path] if path else []


def map_category(icon: str) -> str:
    key = icon.strip().lower()
    return ICON_TO_CATEGORY.get(key, key or 'pin')


def convert_row(row: tuple, idx: dict[str, int], row_number: int) -> dict | None:
    lat = row[idx['lat']]
    lon = row[idx['lon']]
    if lat is None or lon is None:
        return None

    name = clean_text(row[idx['name']])
    if not name:
        return None

    excel_icon = clean_text(row[idx['icon']]) or 'pin'
    category = map_category(excel_icon)
    site_code = clean_text(row[idx['event_site_code']])
    point_id = site_code or f'mia-{slugify(name)}-{row_number}'

    point: dict = {
        'id': point_id,
        'name': name,
        'type': [category],
        'description': clean_text(row[idx['description']]),
        'icon': excel_icon,
        'picture': to_picture(row[idx['img']]),
        'link': clean_text(row[idx['link']]),
        'coordinates': {
            'lat': float(lat),
            'lng': float(lon),
        },
    }

    if excel_icon == 'fanfestival':
        point['isCollapsible'] = 'FF.json'

    return point


def convert_sheet(
    xlsx_path: Path,
    sheet_name: str = '20260624',
    *,
    city_filter: str | None = 'mia',
) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet {sheet_name!r} not found. Available: {wb.sheetnames}')

    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    headers = [clean_text(h) for h in next(row_iter)]
    idx = {h: i for i, h in enumerate(headers)}

    required = ['name', 'description', 'img', 'link', 'icon', 'lat', 'lon']
    missing = [col for col in required if col not in idx]
    if missing:
        raise ValueError(f'Missing columns: {missing}')

    points: list[dict] = []
    seen_ids: set[str] = set()

    for row_number, row in enumerate(row_iter, start=2):
        if city_filter and city_filter in idx and not as_bool(row[idx[city_filter]]):
            continue

        point = convert_row(row, idx, row_number)
        if not point:
            continue

        base_id = point['id']
        unique_id = base_id
        suffix = 2
        while unique_id in seen_ids:
            unique_id = f'{base_id}-{suffix}'
            suffix += 1
        point['id'] = unique_id
        seen_ids.add(unique_id)
        points.append(point)

    return points


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT
    city_filter = sys.argv[3] if len(sys.argv) > 3 else 'mia'
    city_filter = None if city_filter.lower() == 'all' else city_filter

    points = convert_sheet(xlsx_path, city_filter=city_filter)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(points, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')

    print(f'Wrote {len(points)} points to {out_path}')
    if city_filter:
        print(f'Filtered by column {city_filter!r}=True')


if __name__ == '__main__':
    main()
